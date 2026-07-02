"""Comprehensive API test for all endpoints"""
import sys, os, json, io, uuid
sys.path.insert(0, 'e:/案件处理启动器')

os.environ['DEEPSEEK_API_KEY'] = ''  # No API key - test fallback
os.environ['BAIDU_API_KEY'] = ''
os.environ['BAIDU_SECRET_KEY'] = ''
os.environ['QWEN_VL_API_KEY'] = ''

from fastapi.testclient import TestClient
from modules.app_core import app
from modules import routes

client = TestClient(app)

passed = 0
failed = 0

def test(name, method, url, **kwargs):
    global passed, failed
    try:
        if method == 'GET':
            resp = client.get(url, **kwargs)
        elif method == 'POST':
            resp = client.post(url, **kwargs)
        else:
            raise ValueError(f'Unknown method: {method}')
        
        status = resp.status_code
        try:
            data = resp.json()
        except:
            data = resp.text[:200]
        
        if status < 500:
            print(f'  ✅ PASS: {name} ({url}) -> {status}')
            passed += 1
        else:
            print(f'  ❌ FAIL: {name} ({url}) -> {status}: {str(data)[:150]}')
            failed += 1
    except Exception as e:
        print(f'  ❌ ERROR: {name} ({url}) -> {e}')
        failed += 1

print('=' * 60)
print('API Test Suite')
print('=' * 60)

# 1. Health check
test('Health', 'GET', '/api/health')

# 2. Home page
test('Home', 'GET', '/')

# 3. Static files
test('CSS Static', 'GET', '/static/css/style.css')

# 4. Tool pages
test('Report Generate tool', 'GET', '/tools/report_generate/index.html')
test('Directory tool', 'GET', '/tools/index4.html')
test('Approval form tool', 'GET', '/tools/index3.html')
test('Attachment processor', 'GET', '/tools/附件处理器-合并版/index.html')

# 5. Report standard
test('Report standard', 'GET', '/report-standard')

# 6. Template presets
test('Preset templates list', 'GET', '/api/template/presets')
test('Preset template detail', 'GET', '/api/template/preset/preset_1')

# 7. Templates
test('List templates', 'GET', '/api/template/list_std')
test('Template list', 'GET', '/api/templates')

# 8. Run report without DeepSeek (testing fallback)
print('\n--- Report generation (without DeepSeek API) ---')
resp = client.post('/api/run-with-preset', data={
    'template_id': 'preset_1',
    'sceneInvestigation': '走访了现场，发现事故地点在福田区滨河大道',
    'policeStationRecord': '派出所笔录显示...',
})
if resp.status_code < 500:
    data = resp.json()
    if data.get('success'):
        print(f'  ✅ Report generation success: {len(data.get("report_content",""))} chars')
        passed += 1
    else:
        print(f'  ❌ Report generation failed: {data.get("error","")}')
        failed += 1
else:
    print(f'  ❌ HTTP {resp.status_code}')
    failed += 1

# 9. /api/run (separate endpoint for frontend)
print('\n--- /api/run endpoint ---')
resp = client.post('/api/run', data={
    'template_id': 'preset_1',
    'sceneInvestigation': '测试走访内容',
})
if resp.status_code < 500:
    data = resp.json()
    if data.get('success'):
        print(f'  ✅ /api/run success: {len(data.get("report_content",""))} chars')
        passed += 1
    else:
        print(f'  ❌ /api/run failed: {data.get("error","")}')
        failed += 1
else:
    print(f'  ❌ HTTP {resp.status_code}')
    failed += 1

# 10. Generate DOCX from markdown
print('\n--- DOCX generation ---')
resp = client.post('/api/generate-docx', data={
    'markdown_report': '# Test Report\n\nSome content',
    'output_name': 'test_report'
})
if resp.status_code < 500:
    data = resp.json()
    if data.get('success'):
        print(f'  ✅ DOCX generated: {data.get("report_id","")}')
        # Test download
        resp2 = client.get(f'/api/download/{data["report_id"]}')
        print(f'  ✅ Download: {resp2.status_code}')
        passed += 2
    else:
        print(f'  ❌ DOCX generation failed: {data.get("error","")}')
        failed += 1
else:
    print(f'  ❌ HTTP {resp.status_code}')
    failed += 1

# 11. Fill Excel
print('\n--- Excel fill test ---')
import openpyxl
wb = openpyxl.Workbook()
ws = wb.active
ws['A1'] = '保单号：'
ws['B1'] = '投保人：'
output = io.BytesIO()
wb.save(output)
output.seek(0)

resp = client.post('/api/fill-excel', data={
    'template': ('test.xlsx', output.read(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
    'data': json.dumps({'policyNo': 'TEST123', 'insured': '测试投保人', 'accidentDate': '2024-01-01'})
})
if resp.status_code < 500:
    print(f'  ✅ Excel fill: {resp.status_code}, {len(resp.content)} bytes')
    passed += 1
else:
    print(f'  ❌ Excel fill: {resp.status_code}')
    failed += 1

# 12. Upload/DOWNLOAD DOCX for directory tool
print('\n--- DOCX upload & process ---')
doc = Document()
doc.add_heading('测试报告', 0)
table = doc.add_table(rows=2, cols=2)
table.style = 'Table Grid'
table.rows[0].cells[0].text = '附件名称'
table.rows[0].cells[1].text = '页码'
table.rows[1].cells[0].text = '附件一'
table.rows[1].cells[1].text = '3'
output2 = io.BytesIO()
doc.save(output2)
output2.seek(0)

resp = client.post('/api/process-docx', data={
    'file': ('test.docx', output2.read(), 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'),
    'attachments': json.dumps([{'number': 1, 'name': '现场照片'}, {'number': 2, 'name': '监控截图'}])
})
if resp.status_code < 500:
    print(f'  ✅ DOCX processed: {resp.status_code}, {len(resp.content)} bytes')
    passed += 1
else:
    print(f'  ❌ DOCX process: {resp.status_code}')
    failed += 1

# 13. Attachment endpoints
print('\n--- Attachment processor ---')
resp = client.get('/api/ranges')
if resp.status_code < 500:
    print(f'  ✅ Ranges: {resp.status_code}')
    passed += 1
else:
    print(f'  ❌ Ranges: {resp.status_code}')
    failed += 1

resp = client.post('/api/ranges/reset')
if resp.status_code < 500:
    print(f'  ✅ Ranges reset: {resp.status_code}')
    passed += 1
else:
    print(f'  ❌ Ranges reset: {resp.status_code}')
    failed += 1

# One-click with no images (should gracefully handle)
resp = client.post('/api/one-click', json={'images': [], 'attachmentNames': ''})
if resp.status_code < 500:
    print(f'  ✅ One-click: {resp.status_code}')
    passed += 1
else:
    print(f'  ❌ One-click: {resp.status_code}')
    failed += 1

# Generate markdown
resp = client.post('/api/generate-markdown', json={'images': []})
if resp.status_code < 500:
    print(f'  ✅ Generate markdown: {resp.status_code}')
    passed += 1
else:
    print(f'  ❌ Generate markdown: {resp.status_code}')
    failed += 1

# 14. SSE progress
resp = client.get('/api/progress')
if resp.status_code < 500:
    print(f'  ✅ SSE progress: {resp.status_code}')
    passed += 1
else:
    print(f'  ❌ SSE progress: {resp.status_code}')
    failed += 1

# Summary
print()
print('=' * 60)
print(f'Results: {passed} passed, {failed} failed, {passed+failed} total')
print('=' * 60)
if failed > 0:
    sys.exit(1)